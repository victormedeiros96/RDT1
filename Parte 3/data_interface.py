import pandas as pd
import json
from sqlalchemy import create_engine, text
from typing import Any, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Side,Font
import os

class DataSourceInterface:
    def __init__(self, json_path: Optional[str] = None, postgres_url: Optional[str] = None, table_name: Optional[str] = None):
        self.json_path = json_path
        self.postgres_url = postgres_url
        self.table_name = table_name
        self.engine = create_engine(postgres_url) if postgres_url else None

    def read_json(self) -> Any:
        if not self.json_path:
            raise ValueError("Caminho do arquivo JSON não definido.")
        with open(self.json_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def write_json(self, data: Any):
        if not self.json_path:
            raise ValueError("Caminho do arquivo JSON não definido.")
        with open(self.json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def read_postgres(self) -> Any:
        if not self.engine or not self.table_name:
            raise ValueError("Configuração do Postgres não definida.")
        with self.engine.connect() as conn:
            result = conn.execute(text(f"SELECT * FROM {self.table_name}"))
            return [dict(row) for row in result]

    def write_postgres(self, data: Dict):
        if not self.engine or not self.table_name:
            raise ValueError("Configuração do Postgres não definida.")
        with self.engine.connect() as conn:
            conn.execute(text(f"INSERT INTO {self.table_name} (data) VALUES (:data)"), {"data": json.dumps(data)})

    def read_excel(self, excel_path: str) -> pd.DataFrame:
        return pd.read_excel(excel_path)

    def write_excel(self, df: pd.DataFrame, excel_path: str):
        df.to_excel(excel_path, index=False)

    def createRange(self, ws, row: int, step: float):
       
        ws.merge_cells(start_row=row, start_column=4, end_row=row+step, end_column=4)
        cell_faixa = ws.cell(row=row, column=4)
        cell_faixa.value = "3E"

        ws.merge_cells(start_row=row+step+1, start_column=4, end_row=row+2*step+1, end_column=4)
        cell_faixa = ws.cell(row=row+step+1, column=4)
        cell_faixa.value = "1E"

        ws.merge_cells(start_row=row+2*step+2, start_column=4, end_row=row+3*step+2, end_column=4)
        cell_faixa = ws.cell(row=row+2*step+2, column=4)
        cell_faixa.value = "1D"

        ws.merge_cells(start_row=row+3*step+3, start_column=4, end_row=row+4*step+3, end_column=4)
        cell_faixa = ws.cell(row=row+3*step+3, column=4)
        cell_faixa.value = "3D"
    
    def write_excel_with_title(self, excel_path: str, titulo: str,km:int, km_final:int=1,df: pd.DataFrame=None):
       
        wb = Workbook()
        ws = wb.active
   

        for km_count in range(km_final-km):
            df_images= df[df['KM'] == km_count].reset_index(drop=True)
           
            step=59*km_count
            # Linha 7: merge colunas A, B e C e coloca o texto "KM estaca"
            ws.merge_cells(start_row=7+step, start_column=1, end_row=7+step, end_column=4)
            ws.row_dimensions[7+step].height = 25  # altura em pontos
            ws.column_dimensions["A"].width  = 5
            ws.column_dimensions["b"].width  = 5  
            ws.column_dimensions["c"].width  = 5  
            ws.column_dimensions["D"].width  = 5
            cell_km = ws.cell(row=7+step, column=1)
            cell_km.value = "KM estaca"
            cell_km.alignment = Alignment(horizontal="center", vertical="center")
           
            # Coluna A: merge da linha 8 a 19 e texto "Defeitos"
            ws.merge_cells(start_row=8+step, start_column=2, end_row=19+step, end_column=2)
            cell_tmp=ws.cell(row=8+step, column=2)
            cell_tmp.border = Border()  # Remove todas as bordas da célula
            ws.merge_cells(start_row=8+step, start_column=1, end_row=19+step, end_column=1)
            cell_defeitos = ws.cell(row=8+step, column=1)
            cell_defeitos.value = "Defeitos"
            cell_defeitos.alignment = Alignment(horizontal="center", vertical="center",text_rotation=90)
            # Coluna C: merge da linha 8 a 19 e texto "Faixa"
            ws.merge_cells(start_row=8+step, start_column=3, end_row=19+step, end_column=3)
            cell_faixa = ws.cell(row=8+step, column=3)
            cell_faixa.value = "Faixa"
            cell_faixa.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
            self.createRange(ws,8+step,2)
        

            #Posicionamento KMs
            for x in range(0,26,1): #26
                ws.merge_cells(start_row=7+step, start_column=5+(x*40), end_row=7+step, end_column=44+(x*40))
                for col in range(5+(x*40), 45+(x*40)):
                    col_letter = ws.cell(row=1, column=col).column_letter
                    ws.column_dimensions[col_letter].width = 0.25
               
                cell_faixa = ws.cell(row=7+step, column=5+(x*40))
                cell_faixa.value = km+km_count+(x*0.04)
                borda_superior_inferior = Border(
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                        left=Side(style=None),
                        right=Side(style=None)
                    )

                cell_faixa.border = borda_superior_inferior
            #Pintar as células com as imagens
             
            for quadrante in range(0, 49,1):
               
                df_temp= df_images[df_images['quadrante'] == quadrante]
             
                for index, row in df_temp.iterrows():
                    col_num=int(row['linha'])+quadrante*20+5
                    row_num=6-int(row['coluna'])+11+step
                   
                    color= self.obter_cor_por_classe(row['classe']) 
                    cell = ws.cell(row=row_num, column=col_num)
                    self.pintar_celula(ws, row=row_num, col=col_num, cor=color)
                    image=row['arquivo_imagem'].split("/")[4]
                    #cell.hyperlink= f"file:///{os.path.abspath(row['arquivo_imagem'])}"
                    cell.hyperlink= f"file:///D:/Fotos/{image}"
                    cell.value=" "
                 
            legendas_soma= df_images.groupby('classe')["area"].sum()
           
            #IRI
            ws.merge_cells(start_row=20+step, start_column=1, end_row=20+step, end_column= ws.max_column)
            ws.merge_cells(start_row=21+step, start_column=1, end_row=24+step, end_column=3)
            self.createRange(ws,21+step,0)
            cell_defeitos = ws.cell(row=21+step, column=1)
            cell_defeitos.value = "IRI\n(m/km)"
            cell_defeitos.alignment = Alignment(horizontal="center", vertical="center")

            #Degrau
            ws.merge_cells(start_row=25+step, start_column=1, end_row=25+step, end_column= ws.max_column)
            ws.merge_cells(start_row=26+step, start_column=1, end_row=26+step, end_column=4)
            cell_degraus = ws.cell(row=26+step, column=1)
            cell_degraus.value = "Degrau do Ac. (mm)"
            cell_defeitos.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=27+step, start_column=1, end_row=27+step, end_column= ws.max_column)


            #ATR
            ws.merge_cells(start_row=28+step, start_column=1, end_row=35+step, end_column=3)
            self.createRange(ws,28+step,1)
            cell_defeitos = ws.cell(row=28+step, column=1)
            cell_defeitos.value = "ATR\n(mm)"
            cell_defeitos.alignment = Alignment(horizontal="center", vertical="center")


            #degrau 2
            ws.merge_cells(start_row=36+step, start_column=1, end_row=36+step, end_column= ws.max_column)
            ws.merge_cells(start_row=37+step, start_column=1, end_row=37+step, end_column=4)
            cell_degraus = ws.cell(row=37+step, column=1)
            cell_degraus.value = "Degrau do Ac. (mm)"
            cell_defeitos.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=38+step, start_column=1, end_row=38+step, end_column= ws.max_column)

            #OBS
            ws.merge_cells(start_row=39+step, start_column=1, end_row=46+step, end_column=3)
            self.createRange(ws,39+step,1)
            cell_defeitos = ws.cell(row=39+step, column=1)
            cell_defeitos.value = "OBS"
            cell_defeitos.alignment = Alignment(horizontal="center", vertical="center")


            #Legendas 
            ws.merge_cells(start_row=47+step, start_column=1, end_row=47+step, end_column= ws.max_column)
            ws.merge_cells(start_row=48+step, start_column=1, end_row=48+step, end_column= ws.max_column)
            cell_title_legend= ws.cell(row=48+step,column=1)
            cell_title_legend.value='Legendas'
            cell_title_legend.alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="left",readingOrder=4)
            cell_title_legend.font= Font(bold=True, size=16)
            # ws.merge_cells(start_row=47+step, start_column=1, end_row=60+step, end_column= ws.max_column)
            ws.cell(row=49+step, column=15).value = "Couro de Jacaré:"
            ws.cell(row=49+step, column=15).alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)
            ws.cell(row=49+step, column=15).font= Font(bold=True, size=12)
            color_jacare= self.obter_cor_por_classe("Couro de Jacaré")
            
            self.pintar_celula(ws,row=49+step, col=16,cor=color_jacare)
            self.pintar_celula(ws,row=49+step, col=17,cor=color_jacare)
            self.pintar_celula(ws,row=49+step, col=18,cor=color_jacare)
            ws.cell(row=49+step, column=19).value =f"{legendas_soma["Couro de Jacaré"]/1000000:.2f} m²".replace('.', ',')
            ws.cell(row=49+step, column=19).alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)

            ws.cell(row=51+step, column=15).value = "Trincas:"
            ws.cell(row=51+step, column=15).font= Font(bold=True, size=12)
            ws.cell(row=51+step, column=15).alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)
            color_trincas= self.obter_cor_por_classe("Trincas")
            self.pintar_celula(ws, row=51+step, col=16, cor=color_trincas)
            self.pintar_celula(ws, row=51+step, col=17, cor=color_trincas)
            self.pintar_celula(ws, row=51+step, col=18, cor=color_trincas)
            ws.cell(row=51+step, column=19).value =f"{legendas_soma["Trincas"]/1000000:.2f} m²".replace('.', ',')
            ws.cell(row=51+step, column=19).font= Font(bold=False, size=12)
            ws.cell(row=51+step, column=19).alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)
            
            ws.cell(row=49+step, column=210).value = "Remendo:"
            ws.cell(row=49+step, column=210).font= Font(bold=True, size=12)
            ws.cell(row=49+step, column=210).alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)
            color_remendo= self.obter_cor_por_classe("Remendo") 
            self.pintar_celula(ws, row=49+step, col=211, cor=color_remendo)
            self.pintar_celula(ws, row=49+step, col=212, cor=color_remendo)
            self.pintar_celula(ws, row=49+step, col=213, cor=color_remendo)
            remendo=ws.cell(row=49+step, column=214)
            remendo.value =  f"{legendas_soma["remendo"]/1000000:.2f} m²".replace('.', ',')
            remendo.font= Font(bold=False, size=12)
            remendo.alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)

            ws.cell(row=51+step, column=210).value = "Panela"
            ws.cell(row=51+step, column=210).alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)
            ws.cell(row=51+step, column=210).font= Font(bold=True, size=12)
            color_panela= self.obter_cor_por_classe("Panela") 
            self.pintar_celula(ws, row=51+step, col=211, cor=color_panela)
            self.pintar_celula(ws, row=51+step, col=212, cor=color_panela)
            self.pintar_celula(ws, row=51+step, col=213, cor=color_panela)
            ws.cell(row=51+step, column=214).value =  f"{legendas_soma["panela"]/1000000:.2f} m²".replace('.', ',')
            ws.cell(row=51+step, column=214).font= Font(bold=False, size=12)
            ws.cell(row=51+step, column=214).alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)


            #Dados iniciais
            ws.merge_cells(start_row=2+step, start_column=1, end_row=2+step, end_column= ws.max_column)
            ws.cell(row=2+step, column=1).value = ""
            self.pintar_bordas_brancos(ws, row=2+step, col_inicio=1, col_fim=ws.max_column)
            ws.merge_cells(start_row=5+step, start_column=1, end_row=5+step, end_column= ws.max_column)
            ws.merge_cells(start_row=6+step, start_column=1, end_row=6+step, end_column= ws.max_column)
            ws.cell(row=6+step, column=1).value = ""
            ws.cell(row=5+step, column=1).value = ""
            ws.row_dimensions[3+step].height = 35
            ws.row_dimensions[4+step].height = 35
            self.remover_bordas_internas_linha(ws,row=3+step,col_inicio=1,col_fim=472)
            #Info de Rodovia
            cell_rodovia=ws.cell(row=3+step, column=35)
            cell_rodovia.value = "Rodovia:"
            cell_rodovia.font = Font(bold=True, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False,  vertical="center",horizontal="right",readingOrder=4)
            
            cell_rodovia=ws.cell(row=3+step, column=36)
            cell_rodovia.value = "MT-140 MT"
            cell_rodovia.font = Font(bold=False, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False,  vertical="center",readingOrder=4)

            cell_rodovia=ws.cell(row=3+step, column=221)
            cell_rodovia.value = "Data:"
            cell_rodovia.font = Font(bold=True, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)

            cell_rodovia=ws.cell(row=3+step, column=222)
            cell_rodovia.value = "27/06/2025"
            cell_rodovia.font = Font(bold=False, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)

            cell_rodovia=ws.cell(row=3+step, column=443)
            cell_rodovia.value = "Km Inicial:"
            cell_rodovia.font = Font(bold=True, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)

            cell_rodovia=ws.cell(row=3+step, column=444)
            cell_rodovia.value = str(km+km_count)
            cell_rodovia.font = Font(bold=False, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)

            cell_rodovia=ws.cell(row=4+step, column=443)
            cell_rodovia.value = "Km Final:"
            cell_rodovia.font = Font(bold=True, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False, vertical="center",horizontal="right",readingOrder=4)

            cell_rodovia=ws.cell(row=4+step, column=444)
            cell_rodovia.value = str(km+km_count+1)
            cell_rodovia.font = Font(bold=False, size=14)
            cell_rodovia.alignment = Alignment(shrink_to_fit=False, vertical="center",readingOrder=4)
            self.pintar_bordas_brancos(ws, row=3+step, col_inicio=1, col_fim=ws.max_column,colorUpper="575655")
            self.pintar_bordas_brancos(ws, row=4+step, col_inicio=1, col_fim=ws.max_column,colorBottom="575655")
            self.pintar_bordas_brancos(ws, row=5+step, col_inicio=1, col_fim=ws.max_column)
            self.pintar_bordas_brancos(ws, row=6+step, col_inicio=1, col_fim=ws.max_column,colorBottom="575655")
            self.pintar_bordas_brancos(ws, row=7+step, col_inicio=1, col_fim=ws.max_column,colorBottom="575655",colorLeft="575655")
            self.pintar_bordas_brancos(ws, row=48+step, col_inicio=1, col_fim=ws.max_column,colorUpper="575655")
            self.pintar_bordas_brancos(ws, row=49+step, col_inicio=1, col_fim=ws.max_column,row_fim=61+step)
            self.pintar_bordas_brancos(ws, row=49+step, col_inicio=16, col_fim=18,colorBottom=color_jacare,colorUpper=color_jacare,colorLeft=color_jacare, colorRight=color_jacare)
            self.pintar_bordas_brancos(ws, row=51+step, col_inicio=16, col_fim=18,colorBottom=color_trincas,colorUpper=color_trincas,colorLeft=color_trincas, colorRight=color_trincas)
            self.pintar_bordas_brancos(ws, row=49+step, col_inicio=211, col_fim=213,colorBottom=color_remendo,colorUpper=color_remendo,colorLeft=color_remendo, colorRight=color_remendo)
            self.pintar_bordas_brancos(ws, row=51+step, col_inicio=211, col_fim=213,colorBottom=color_panela,colorUpper=color_panela,colorLeft=color_panela, colorRight=color_panela)
            
            
            




        ##TITLE
       
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= ws.max_column)
        cell = ws.cell(row=1, column=1)
        cell.value = titulo
        cell.font=Font(bold=True, size=26)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=Side(style="thin"),
            bottom=Side(None),
            left=Side(style="thin"),
            right=Side(style="thin")
        )
        ws.row_dimensions[1].height = 50  # altura em pontos
        wb.save(excel_path)

    def remover_bordas_internas_linha(self, ws, row, col_inicio, col_fim):
        """
        Remove todas as bordas de uma linha (todas as células do intervalo).
        ws: worksheet
        row: número da linha (int)
        col_inicio: coluna inicial (int)
        col_fim: coluna final (int)
        """
        print(row, col_inicio, col_fim)
        for col in range(col_inicio, col_fim + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
            top=Side(None),
            bottom=Side(None),
            left=Side(None),
            right=Side(None)
        )

    def pintar_bordas_brancos(self, ws, row, col_inicio, col_fim,colorUpper='FFFFFF',colorRight='FFFFFF',colorLeft="FFFFFF",colorBottom="FFFFFF",row_fim=None):
        """
        Remove todas as bordas de uma linha (todas as células do intervalo).
        ws: worksheet
        row: número da linha (int)
        col_inicio: coluna inicial (int)
        col_fim: coluna final (int)
        """
        if row_fim==None:
            row_fim=row+1
        borderUpper=Side(border_style='thin', color=colorUpper)
        borderRight=Side(border_style='thin', color=colorRight)
        borderLeft=Side(border_style='thin', color=colorLeft)
        borderBottom=Side(border_style='thin', color=colorBottom)
        print(row, col_inicio, col_fim)
        for r in range(row, row_fim):
            for col in range(col_inicio, col_fim + 1):
                cell = ws.cell(row=r, column=col)
                cell.border =Border(left=borderLeft, 
                            right=borderRight, 
                            top=borderUpper, 
                            bottom=borderBottom)
    
    def pintar_celula(self, ws, row, col, cor="FFFF00"):
        """
        Pinta uma célula com a cor especificada.
        ws: worksheet
        row: número da linha (int)
        col: número da coluna (int)
        cor: código hex da cor (string) - padrão amarelo
        """
        from openpyxl.styles import PatternFill
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
    
    def pintar_intervalo_celulas(self, ws, start_row, start_col, end_row, end_col, cor="FFFF00"):
        """
        Pinta um intervalo de células com a cor especificada.
        """
        from openpyxl.styles import PatternFill
        fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill

    def obter_cor_por_classe(self, classe):
        """
        Retorna uma cor baseada na classe/categoria da detecção.
        classe: string com o nome da classe
        retorna: código hex da cor (string)
        """
        cores_por_classe = {
            # Defeitos de pavimento
            "trincas": "FFFF00",          # Amarelo
            "panela": "FF0000",         # Vermelho

            "couro de jacaré": "00FF00",           # Verde
            "remendo": "0000FF",        # Azul
           
            "default": "FF00FF",  # Branco padrão para classes não mapeadas
        }
        
        # Converte para minúsculo para evitar problemas de case
        classe_lower = classe.lower() if classe else ""
    
        return cores_por_classe.get(classe_lower, cores_por_classe["default"])
